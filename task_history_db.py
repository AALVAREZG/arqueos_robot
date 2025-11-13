"""
Task History Database Manager
Handles SQLite persistence for task execution history
"""

import sqlite3
import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Any
from contextlib import contextmanager
from dataclasses import asdict

logger = logging.getLogger(__name__)


class TaskHistoryDB:
    """Manages SQLite database for task history persistence"""

    def __init__(self, db_path: str = "arqueos_history.db"):
        """
        Initialize database connection

        Args:
            db_path: Path to SQLite database file
        """
        self.db_path = Path(db_path)
        self._init_database()

    @contextmanager
    def _get_connection(self):
        """Context manager for database connections"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row  # Return rows as dictionaries
        try:
            yield conn
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.error(f"Database error: {e}")
            raise
        finally:
            conn.close()

    def _init_database(self):
        """Create tables if they don't exist"""
        with self._get_connection() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS task_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    task_id TEXT UNIQUE NOT NULL,
                    operation_number TEXT,
                    amount REAL,
                    date TEXT,
                    cash_register TEXT,
                    third_party TEXT,
                    nature TEXT,
                    description TEXT,
                    total_line_items INTEGER,
                    status TEXT,
                    started_at TIMESTAMP,
                    completed_at TIMESTAMP,
                    duration_seconds REAL,
                    error_message TEXT,
                    raw_data TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Create indexes for common queries
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_task_id
                ON task_history(task_id)
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_completed_at
                ON task_history(completed_at DESC)
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_status
                ON task_history(status)
            """)

        logger.info(f"Task history database initialized: {self.db_path}")

    def save_task(self, task_data: Dict[str, Any]) -> bool:
        """
        Save or update a task record

        Args:
            task_data: Dictionary with task information

        Returns:
            True if successful, False otherwise
        """
        try:
            with self._get_connection() as conn:
                conn.execute("""
                    INSERT OR REPLACE INTO task_history (
                        task_id, operation_number, amount, date, cash_register,
                        third_party, nature, description, total_line_items,
                        status, started_at, completed_at, duration_seconds,
                        error_message, raw_data
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    task_data.get('task_id'),
                    task_data.get('operation_number'),
                    task_data.get('amount'),
                    task_data.get('date'),
                    task_data.get('cash_register'),
                    task_data.get('third_party'),
                    task_data.get('nature'),
                    task_data.get('description'),
                    task_data.get('total_line_items'),
                    task_data.get('status'),
                    task_data.get('started_at'),
                    task_data.get('completed_at'),
                    task_data.get('duration_seconds'),
                    task_data.get('error_message'),
                    json.dumps(task_data.get('raw_data', {}))
                ))
            return True
        except Exception as e:
            logger.error(f"Failed to save task {task_data.get('task_id')}: {e}")
            return False

    def update_task_status(self, task_id: str, status: str,
                          completed_at: Optional[datetime] = None,
                          duration_seconds: Optional[float] = None,
                          error_message: Optional[str] = None) -> bool:
        """
        Update task status and completion info

        Args:
            task_id: Task identifier
            status: New status ('completed', 'failed', 'error')
            completed_at: Completion timestamp
            duration_seconds: Task duration
            error_message: Error message if failed

        Returns:
            True if successful
        """
        try:
            with self._get_connection() as conn:
                conn.execute("""
                    UPDATE task_history
                    SET status = ?,
                        completed_at = ?,
                        duration_seconds = ?,
                        error_message = ?
                    WHERE task_id = ?
                """, (status, completed_at, duration_seconds, error_message, task_id))
            return True
        except Exception as e:
            logger.error(f"Failed to update task {task_id}: {e}")
            return False

    def get_all_tasks(self, limit: Optional[int] = None,
                      status_filter: Optional[str] = None,
                      order_by: str = "completed_at DESC") -> List[Dict]:
        """
        Retrieve all tasks from history

        Args:
            limit: Maximum number of records to return
            status_filter: Filter by status ('completed', 'failed', 'error')
            order_by: SQL ORDER BY clause

        Returns:
            List of task dictionaries
        """
        try:
            with self._get_connection() as conn:
                query = "SELECT * FROM task_history"
                params = []

                if status_filter:
                    query += " WHERE status = ?"
                    params.append(status_filter)

                query += f" ORDER BY {order_by}"

                if limit:
                    query += " LIMIT ?"
                    params.append(limit)

                cursor = conn.execute(query, params)
                rows = cursor.fetchall()

                # Convert to list of dicts
                return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"Failed to retrieve tasks: {e}")
            return []

    def get_task_by_id(self, task_id: str) -> Optional[Dict]:
        """Get a specific task by ID"""
        try:
            with self._get_connection() as conn:
                cursor = conn.execute(
                    "SELECT * FROM task_history WHERE task_id = ?",
                    (task_id,)
                )
                row = cursor.fetchone()
                return dict(row) if row else None
        except Exception as e:
            logger.error(f"Failed to retrieve task {task_id}: {e}")
            return None

    def get_statistics(self) -> Dict[str, Any]:
        """Get overall statistics from history"""
        try:
            with self._get_connection() as conn:
                cursor = conn.execute("""
                    SELECT
                        COUNT(*) as total_tasks,
                        SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed,
                        SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) as failed,
                        SUM(CASE WHEN status = 'error' THEN 1 ELSE 0 END) as errors,
                        AVG(duration_seconds) as avg_duration,
                        MIN(started_at) as first_task,
                        MAX(completed_at) as last_task
                    FROM task_history
                """)
                row = cursor.fetchone()
                return dict(row) if row else {}
        except Exception as e:
            logger.error(f"Failed to retrieve statistics: {e}")
            return {}

    def search_tasks(self, search_term: str, limit: int = 100) -> List[Dict]:
        """
        Search tasks by operation number, description, or third party

        Args:
            search_term: Search string
            limit: Max results

        Returns:
            List of matching tasks
        """
        try:
            with self._get_connection() as conn:
                cursor = conn.execute("""
                    SELECT * FROM task_history
                    WHERE task_id LIKE ?
                       OR operation_number LIKE ?
                       OR description LIKE ?
                       OR third_party LIKE ?
                    ORDER BY completed_at DESC
                    LIMIT ?
                """, (f"%{search_term}%", f"%{search_term}%",
                      f"%{search_term}%", f"%{search_term}%", limit))

                rows = cursor.fetchall()
                return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"Search failed: {e}")
            return []

    def export_to_json(self, filepath: str, limit: Optional[int] = None) -> bool:
        """Export task history to JSON file"""
        try:
            tasks = self.get_all_tasks(limit=limit)
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(tasks, f, indent=2, default=str)
            logger.info(f"Exported {len(tasks)} tasks to {filepath}")
            return True
        except Exception as e:
            logger.error(f"JSON export failed: {e}")
            return False

    def export_to_csv(self, filepath: str, limit: Optional[int] = None) -> bool:
        """Export task history to CSV file"""
        try:
            import csv
            tasks = self.get_all_tasks(limit=limit)

            if not tasks:
                return False

            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                # Use all keys from first task as headers
                fieldnames = list(tasks[0].keys())
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(tasks)

            logger.info(f"Exported {len(tasks)} tasks to {filepath}")
            return True
        except Exception as e:
            logger.error(f"CSV export failed: {e}")
            return False

    def export_to_excel(self, filepath: str, limit: Optional[int] = None) -> bool:
        """Export task history to Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            tasks = self.get_all_tasks(limit=limit)

            if not tasks:
                return False

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Task History"

            # Headers
            headers = list(tasks[0].keys())
            ws.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Data rows
            for task in tasks:
                ws.append([task.get(h) for h in headers])

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(filepath)
            logger.info(f"Exported {len(tasks)} tasks to {filepath}")
            return True
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            return False
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return False

    def clear_history(self, before_date: Optional[datetime] = None) -> int:
        """
        Clear task history (use with caution!)

        Args:
            before_date: If provided, only delete tasks before this date

        Returns:
            Number of deleted records
        """
        try:
            with self._get_connection() as conn:
                if before_date:
                    cursor = conn.execute(
                        "DELETE FROM task_history WHERE completed_at < ?",
                        (before_date,)
                    )
                else:
                    cursor = conn.execute("DELETE FROM task_history")

                deleted = cursor.rowcount
                logger.warning(f"Deleted {deleted} task records from history")
                return deleted
        except Exception as e:
            logger.error(f"Failed to clear history: {e}")
            return 0


# Global singleton instance
_task_history_db = None

def get_task_history_db() -> TaskHistoryDB:
    """Get or create the global TaskHistoryDB instance"""
    global _task_history_db
    if _task_history_db is None:
        _task_history_db = TaskHistoryDB()
    return _task_history_db
